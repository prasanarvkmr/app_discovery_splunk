"""  
Reporting Module
================

This module handles the generation of discovery reports in various formats:
- CSV files
- JSON files
- Excel workbooks (with formatting)
- HTML reports

It also provides utilities for creating visualizations and summary statistics.
Optimized for performance with lazy evaluation and efficient data processing.

Usage:
    from reporting import ReportGenerator
    
    reporter = ReportGenerator(discovered_apps, config)
    reporter.generate_csv("output/applications.csv")
    reporter.generate_excel("output/applications.xlsx")
    reporter.generate_html_report("output/report.html")
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Generator, Iterator
from itertools import islice

from .discovery_strategies import DiscoveredApp
from .config import DiscoveryConfig

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates reports from discovery results.
    
    This class takes discovered applications and generates various
    report formats suitable for different audiences and purposes.
    
    Attributes:
        apps (List[DiscoveredApp]): List of discovered applications
        config (DiscoveryConfig): Configuration settings
        stats (Dict): Discovery statistics
        
    Example:
        >>> reporter = ReportGenerator(discovered_apps, config, stats)
        >>> reporter.generate_csv("applications.csv")
        >>> reporter.generate_excel("applications.xlsx")
    """
    
    def __init__(
        self, 
        apps: List[DiscoveredApp],
        config: DiscoveryConfig,
        stats: Optional[Dict] = None
    ):
        """
        Initialize the report generator.
        
        Args:
            apps: List of discovered applications
            config: Discovery configuration
            stats: Discovery statistics dictionary
        """
        self.apps = apps
        self.config = config
        self.stats = stats or {}
        
        # Ensure output directory exists
        Path(config.output_path).mkdir(parents=True, exist_ok=True)
    
    def _iter_app_rows(self) -> Generator[Dict[str, Any], None, None]:
        """
        Generator that yields app data rows for efficient memory usage.
        
        Yields:
            Dictionary with app data
        """
        for app in self.apps:
            yield {
                "Application": app.name,
                "Confidence Score": app.confidence_score,
                "Event Count": app.event_count,
                "Source Count": len(app.sources),
                "Host Count": len(app.hosts),
                "Sourcetype Count": len(app.sourcetypes),
                "Discovery Methods": ", ".join(app.discovery_methods),
                "Sample Sources": "; ".join(islice(app.sources, 3)),
                "Sample Hosts": "; ".join(islice(app.hosts, 3)),
                "Sourcetypes": "; ".join(app.sourcetypes),
            }
    
    def generate_csv(self, filename: Optional[str] = None) -> str:
        """
        Generate a CSV report of discovered applications.
        Uses streaming writes for better memory efficiency.
        
        Args:
            filename: Output filename (default: applications_TIMESTAMP.csv)
            
        Returns:
            Path to generated file
        """
        import csv
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"applications_{timestamp}.csv"
        
        filepath = Path(self.config.output_path) / filename
        
        # Stream write directly to file
        headers = [
            "Application", "Confidence Score", "Event Count",
            "Source Count", "Host Count", "Sourcetype Count",
            "Discovery Methods", "Sample Sources", "Sample Hosts",
            "Sourcetypes"
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            
            # Use generator for memory efficiency
            for row_data in self._iter_app_rows():
                writer.writerow(row_data)
        
        logger.info(f"CSV report generated: {filepath}")
        return str(filepath)
    
    def _generate_csv_manual(self, filename: Optional[str] = None) -> str:
        """Generate CSV without pandas."""
        import csv
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"applications_{timestamp}.csv"
        
        filepath = Path(self.config.output_path) / filename
        
        headers = [
            "Application", "Confidence Score", "Event Count",
            "Source Count", "Host Count", "Sourcetype Count",
            "Discovery Methods", "Sample Sources", "Sample Hosts",
            "Sourcetypes", "Evidence"
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for app in self.apps:
                writer.writerow([
                    app.name,
                    app.confidence_score,
                    app.event_count,
                    len(app.sources),
                    len(app.hosts),
                    len(app.sourcetypes),
                    "; ".join(app.discovery_methods),
                    "; ".join(list(app.sources)[:3]),
                    "; ".join(list(app.hosts)[:3]),
                    "; ".join(app.sourcetypes),
                    "; ".join(app.evidence[:3])
                ])
        
        logger.info(f"CSV report generated: {filepath}")
        return str(filepath)
    
    def generate_json(self, filename: Optional[str] = None, indent: int = 2) -> str:
        """
        Generate a JSON report of discovered applications.
        Uses efficient serialization for large datasets.
        
        Args:
            filename: Output filename (default: applications_TIMESTAMP.json)
            indent: JSON indentation (None for compact output)
            
        Returns:
            Path to generated file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"applications_{timestamp}.json"
        
        filepath = Path(self.config.output_path) / filename
        
        # Build JSON structure efficiently
        report = {
            "generated_at": datetime.now().isoformat(),
            "config": {
                "splunk_host": self.config.splunk_host,
                "timeframe": self.config.discovery_timeframe,
                "confidence_threshold": self.config.confidence_threshold,
            },
            "statistics": self.stats,
            "summary": self._get_summary_stats(),
            "applications": [
                {
                    "name": app.name,
                    "confidence_score": app.confidence_score,
                    "event_count": app.event_count,
                    "sources": list(islice(app.sources, 100)),  # Limit sources to save space
                    "hosts": list(islice(app.hosts, 50)),  # Limit hosts
                    "sourcetypes": list(app.sourcetypes),
                    "discovery_methods": list(app.discovery_methods),
                    "evidence": app.evidence[:10],  # Limit evidence items
                    "first_seen": app.first_seen,
                    "last_seen": app.last_seen,
                }
                for app in self.apps
            ]
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=indent, default=str)
        
        logger.info(f"JSON report generated: {filepath}")
        return str(filepath)
    
    def _get_summary_stats(self) -> Dict[str, int]:
        """Calculate summary statistics efficiently."""
        high_conf = sum(1 for a in self.apps if a.confidence_score >= 70)
        med_conf = sum(1 for a in self.apps if 40 <= a.confidence_score < 70)
        low_conf = len(self.apps) - high_conf - med_conf
        
        return {
            "total_applications": len(self.apps),
            "high_confidence": high_conf,
            "medium_confidence": med_conf,
            "low_confidence": low_conf,
        }
    
    def generate_excel(self, filename: Optional[str] = None) -> str:
        """
        Generate an Excel report with multiple sheets and formatting.
        
        Args:
            filename: Output filename (default: applications_TIMESTAMP.xlsx)
            
        Returns:
            Path to generated file
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import BarChart, Reference, PieChart
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl pandas"
            )
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"applications_{timestamp}.xlsx"
        
        filepath = Path(self.config.output_path) / filename
        
        # Create workbook
        wb = Workbook()
        
        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Add summary data
        summary_data = [
            ["Splunk Application Discovery Report", ""],
            ["", ""],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Splunk Host", self.config.splunk_host],
            ["Time Range", self.config.discovery_timeframe],
            ["Confidence Threshold", f"{self.config.confidence_threshold}%"],
            ["", ""],
            ["Summary Statistics", ""],
            ["Total Applications Discovered", len(self.apps)],
            ["High Confidence (≥70%)", len([a for a in self.apps if a.confidence_score >= 70])],
            ["Medium Confidence (40-69%)", len([a for a in self.apps if 40 <= a.confidence_score < 70])],
            ["Low Confidence (<40%)", len([a for a in self.apps if a.confidence_score < 40])],
            ["Total Events Analyzed", sum(a.event_count for a in self.apps)],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Style title
        ws_summary["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws_summary.merge_cells("A1:B1")
        
        # ===== Sheet 2: Applications =====
        ws_apps = wb.create_sheet("Applications")
        
        # Headers
        headers = [
            "Application Name", "Confidence", "Event Count",
            "Sources", "Hosts", "Sourcetypes", "Discovery Methods"
        ]
        ws_apps.append(headers)
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws_apps.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for app in self.apps:
            ws_apps.append([
                app.name,
                f"{app.confidence_score}%",
                app.event_count,
                len(app.sources),
                len(app.hosts),
                len(app.sourcetypes),
                ", ".join(app.discovery_methods)
            ])
        
        # Color-code confidence
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in range(2, len(self.apps) + 2):
            confidence_cell = ws_apps.cell(row=row, column=2)
            confidence_value = int(confidence_cell.value.replace("%", ""))
            
            if confidence_value >= 70:
                confidence_cell.fill = green_fill
            elif confidence_value >= 40:
                confidence_cell.fill = yellow_fill
            else:
                confidence_cell.fill = red_fill
        
        # Adjust column widths
        ws_apps.column_dimensions["A"].width = 30
        ws_apps.column_dimensions["B"].width = 12
        ws_apps.column_dimensions["C"].width = 15
        ws_apps.column_dimensions["G"].width = 40
        
        # ===== Sheet 3: Detailed Evidence =====
        ws_evidence = wb.create_sheet("Evidence Details")
        
        evidence_headers = ["Application", "Discovery Method", "Evidence"]
        ws_evidence.append(evidence_headers)
        
        for col, header in enumerate(evidence_headers, 1):
            cell = ws_evidence.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        for app in self.apps:
            for evidence in app.evidence[:5]:  # Limit to 5 evidence items per app
                ws_evidence.append([app.name, "", evidence])
        
        ws_evidence.column_dimensions["A"].width = 25
        ws_evidence.column_dimensions["C"].width = 80
        
        # ===== Sheet 4: Sources =====
        ws_sources = wb.create_sheet("Sources")
        ws_sources.append(["Application", "Source Path"])
        
        for col in range(1, 3):
            cell = ws_sources.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        for app in self.apps:
            for source in list(app.sources)[:10]:  # Limit sources
                ws_sources.append([app.name, source])
        
        ws_sources.column_dimensions["A"].width = 25
        ws_sources.column_dimensions["B"].width = 80
        
        # Save workbook
        wb.save(filepath)
        
        logger.info(f"Excel report generated: {filepath}")
        return str(filepath)
    
    def generate_html_report(self, filename: Optional[str] = None) -> str:
        """
        Generate an HTML report with interactive elements.
        Optimized for performance with efficient string building.
        
        Args:
            filename: Output filename (default: report_TIMESTAMP.html)
            
        Returns:
            Path to generated file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.html"
        
        filepath = Path(self.config.output_path) / filename
        
        # Calculate statistics once (optimized)
        summary = self._get_summary_stats()
        high_conf = summary["high_confidence"]
        med_conf = summary["medium_confidence"]
        low_conf = summary["low_confidence"]
        
        # Generate application rows efficiently
        app_rows_list = []
        for app in self.apps:
            conf_class = "high" if app.confidence_score >= 70 else "medium" if app.confidence_score >= 40 else "low"
            methods = ", ".join(app.discovery_methods)
            
            app_rows_list.append(f"""
            <tr>
                <td>{app.name}</td>
                <td class="confidence {conf_class}">{app.confidence_score}%</td>
                <td>{app.event_count:,}</td>
                <td>{len(app.sources)}</td>
                <td>{len(app.hosts)}</td>
                <td>{methods}</td>
            </tr>
            """)
        
        app_rows = "".join(app_rows_list)  # Single join operation
        
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Splunk Application Discovery Report</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f5f5;
            color: #333;
            line-height: 1.6;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }}
        header {{
            background: linear-gradient(135deg, #1a5276, #2980b9);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
        }}
        header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        header p {{
            opacity: 0.9;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            text-align: center;
        }}
        .stat-card h3 {{
            font-size: 2.5em;
            color: #2980b9;
            margin-bottom: 5px;
        }}
        .stat-card p {{
            color: #666;
            font-size: 0.9em;
        }}
        .stat-card.high h3 {{ color: #27ae60; }}
        .stat-card.medium h3 {{ color: #f39c12; }}
        .stat-card.low h3 {{ color: #e74c3c; }}
        
        .section {{
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 30px;
            overflow: hidden;
        }}
        .section-header {{
            background: #2c3e50;
            color: white;
            padding: 15px 25px;
            font-size: 1.2em;
        }}
        .section-content {{
            padding: 20px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 12px 15px;
            text-align: left;
            border-bottom: 1px solid #eee;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #2c3e50;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .confidence {{
            font-weight: bold;
            padding: 5px 10px;
            border-radius: 20px;
            text-align: center;
        }}
        .confidence.high {{
            background: #d4edda;
            color: #155724;
        }}
        .confidence.medium {{
            background: #fff3cd;
            color: #856404;
        }}
        .confidence.low {{
            background: #f8d7da;
            color: #721c24;
        }}
        
        .search-box {{
            width: 100%;
            padding: 12px 20px;
            border: 2px solid #ddd;
            border-radius: 25px;
            font-size: 1em;
            margin-bottom: 20px;
            outline: none;
        }}
        .search-box:focus {{
            border-color: #2980b9;
        }}
        
        footer {{
            text-align: center;
            padding: 20px;
            color: #666;
        }}
        
        @media (max-width: 768px) {{
            .stats-grid {{
                grid-template-columns: 1fr 1fr;
            }}
            table {{
                font-size: 0.9em;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>🔍 Splunk Application Discovery Report</h1>
            <p>Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | 
               Host: {self.config.splunk_host} | 
               Time Range: {self.config.discovery_timeframe}</p>
        </header>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>{len(self.apps)}</h3>
                <p>Total Applications</p>
            </div>
            <div class="stat-card high">
                <h3>{high_conf}</h3>
                <p>High Confidence (≥70%)</p>
            </div>
            <div class="stat-card medium">
                <h3>{med_conf}</h3>
                <p>Medium Confidence (40-69%)</p>
            </div>
            <div class="stat-card low">
                <h3>{low_conf}</h3>
                <p>Low Confidence (<40%)</p>
            </div>
        </div>
        
        <div class="section">
            <div class="section-header">📋 Discovered Applications</div>
            <div class="section-content">
                <input type="text" class="search-box" id="searchInput" 
                       placeholder="🔍 Search applications..." onkeyup="filterTable()">
                <table id="appTable">
                    <thead>
                        <tr>
                            <th>Application Name</th>
                            <th>Confidence</th>
                            <th>Event Count</th>
                            <th>Sources</th>
                            <th>Hosts</th>
                            <th>Discovery Methods</th>
                        </tr>
                    </thead>
                    <tbody>
                        {app_rows}
                    </tbody>
                </table>
            </div>
        </div>
        
        <footer>
            <p>Splunk Application Discovery Tool | Enterprise Observability</p>
        </footer>
    </div>
    
    <script>
        function filterTable() {{
            const input = document.getElementById('searchInput');
            const filter = input.value.toLowerCase();
            const table = document.getElementById('appTable');
            const rows = table.getElementsByTagName('tr');
            
            for (let i = 1; i < rows.length; i++) {{
                const cells = rows[i].getElementsByTagName('td');
                let found = false;
                
                for (let j = 0; j < cells.length; j++) {{
                    if (cells[j].textContent.toLowerCase().includes(filter)) {{
                        found = true;
                        break;
                    }}
                }}
                
                rows[i].style.display = found ? '' : 'none';
            }}
        }}
    </script>
</body>
</html>
        """
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML report generated: {filepath}")
        return str(filepath)
    
    def generate_all_reports(self, base_name: Optional[str] = None) -> Dict[str, str]:
        """
        Generate all report formats.
        
        Args:
            base_name: Base filename (timestamp added automatically)
            
        Returns:
            Dictionary mapping format to filepath
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = base_name or f"applications_{timestamp}"
        
        reports = {}
        
        reports["csv"] = self.generate_csv(f"{base}.csv")
        reports["json"] = self.generate_json(f"{base}.json")
        reports["html"] = self.generate_html_report(f"{base}.html")
        
        try:
            reports["excel"] = self.generate_excel(f"{base}.xlsx")
        except ImportError:
            logger.warning("Excel report skipped (openpyxl not installed)")
        
        logger.info(f"All reports generated in: {self.config.output_path}")
        return reports
    
    def _to_dataframe(self):
        """Convert applications to pandas DataFrame."""
        import pandas as pd
        
        data = []
        for app in self.apps:
            data.append({
                "Application": app.name,
                "Confidence Score": app.confidence_score,
                "Event Count": app.event_count,
                "Source Count": len(app.sources),
                "Host Count": len(app.hosts),
                "Sourcetype Count": len(app.sourcetypes),
                "Discovery Methods": ", ".join(app.discovery_methods),
                "Sample Sources": "; ".join(list(app.sources)[:3]),
                "Sample Hosts": "; ".join(list(app.hosts)[:3]),
                "Sourcetypes": "; ".join(app.sourcetypes),
            })
        
        return pd.DataFrame(data)


def export_results(
    apps: List[DiscoveredApp],
    config: DiscoveryConfig,
    format: str = "csv",
    filename: Optional[str] = None
) -> str:
    """
    Convenience function to export discovery results.
    
    Args:
        apps: List of discovered applications
        config: Discovery configuration
        format: Output format ('csv', 'json', 'excel', 'html', 'all')
        filename: Optional output filename
        
    Returns:
        Path to generated file(s)
    """
    reporter = ReportGenerator(apps, config)
    
    if format == "csv":
        return reporter.generate_csv(filename)
    elif format == "json":
        return reporter.generate_json(filename)
    elif format == "excel":
        return reporter.generate_excel(filename)
    elif format == "html":
        return reporter.generate_html_report(filename)
    elif format == "all":
        return str(reporter.generate_all_reports())
    else:
        raise ValueError(f"Unknown format: {format}. Use: csv, json, excel, html, all")
